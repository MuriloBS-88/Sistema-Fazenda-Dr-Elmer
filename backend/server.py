from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Literal
import uuid
from datetime import datetime, timezone, date, timedelta
from fastapi.responses import StreamingResponse
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import bcrypt
import jwt

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

JWT_ALGORITHM = "HS256"

def get_jwt_secret():
    return os.environ["JWT_SECRET"]

# ============= AUTH HELPERS =============

def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode("utf-8"), salt)
    return hashed.decode("utf-8")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode("utf-8"), hashed_password.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(hours=24), "type": "access"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Nao autenticado")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Token invalido")
        user_id = payload["sub"]
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario nao encontrado")
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalido")

async def require_admin(request: Request) -> dict:
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Acesso negado. Somente administradores.")
    return user


# ============= AUTH MODELS =============

class LoginInput(BaseModel):
    email: str
    password: str

class RegisterUserInput(BaseModel):
    nome: str
    email: str
    password: str
    role: str = "user"

class UpdateUserInput(BaseModel):
    nome: str
    email: str
    password: Optional[str] = None
    role: str = "user"

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    nome: str
    email: str
    role: str
    created_at: datetime


# ============= AUTH ENDPOINTS =============

@api_router.post("/auth/login")
async def login(input: LoginInput, response: Response):
    email = input.email.lower().strip()
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user or not verify_password(input.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email ou senha incorretos")
    
    access_token = create_access_token(user["id"], user["email"])
    refresh_token = create_refresh_token(user["id"])
    
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    
    return {"id": user["id"], "nome": user["nome"], "email": user["email"], "role": user["role"], "token": access_token}

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    return user

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logout realizado"}

@api_router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="Nao autenticado")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Token invalido")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario nao encontrado")
        access_token = create_access_token(user["id"], user["email"])
        response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=86400, path="/")
        return {"message": "Token renovado"}
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalido")


# ============= USER MANAGEMENT (admin only) =============

@api_router.post("/users", response_model=UserResponse)
async def criar_usuario(input: RegisterUserInput, request: Request):
    admin = await require_admin(request)
    
    email = input.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email ja cadastrado")
    
    user = {
        "id": str(uuid.uuid4()),
        "nome": input.nome,
        "email": email,
        "password_hash": hash_password(input.password),
        "role": input.role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user)
    user.pop("password_hash", None)
    user.pop("_id", None)
    if isinstance(user.get("created_at"), str):
        user["created_at"] = datetime.fromisoformat(user["created_at"])
    return user

@api_router.get("/users", response_model=List[UserResponse])
async def listar_usuarios(request: Request):
    await require_admin(request)
    docs = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    for doc in docs:
        if isinstance(doc.get("created_at"), str):
            doc["created_at"] = datetime.fromisoformat(doc["created_at"])
    return docs

@api_router.delete("/users/{user_id}")
async def deletar_usuario(user_id: str, request: Request):
    admin = await require_admin(request)
    if admin["id"] == user_id:
        raise HTTPException(status_code=400, detail="Voce nao pode deletar a si mesmo")
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    return {"message": "Usuario deletado"}

@api_router.put("/users/{user_id}", response_model=UserResponse)
async def atualizar_usuario(user_id: str, input: UpdateUserInput, request: Request):
    await require_admin(request)
    email = input.email.lower().strip()
    existing = await db.users.find_one({"email": email, "id": {"$ne": user_id}})
    if existing:
        raise HTTPException(status_code=400, detail="Email ja cadastrado por outro usuario")
    update_data = {"nome": input.nome, "email": email, "role": input.role}
    if input.password:
        update_data["password_hash"] = hash_password(input.password)
    result = await db.users.update_one({"id": user_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    doc = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if isinstance(doc.get("created_at"), str):
        doc["created_at"] = datetime.fromisoformat(doc["created_at"])
    return doc


# ============= EXISTING MODELS =============

class CategoriaCreate(BaseModel):
    nome: str
    cor: str = "#4A6741"

class Categoria(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nome: str
    cor: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AnimalCreate(BaseModel):
    tipo: str
    tag: str
    data_nascimento: Optional[date] = None
    peso_atual: Optional[float] = None
    observacoes: Optional[str] = ""

class Animal(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tipo: str
    tag: str
    data_nascimento: Optional[date] = None
    peso_atual: Optional[float] = None
    observacoes: Optional[str] = ""
    status: str = "ativo"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MovimentacaoCreate(BaseModel):
    tipo: Literal["entrada", "saida"]
    motivo: str
    animal_id: Optional[str] = None
    data: date
    valor: Optional[float] = None
    quantidade: int = 1
    tipo_animal: Optional[str] = None
    observacoes: Optional[str] = ""

class Movimentacao(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tipo: Literal["entrada", "saida"]
    motivo: str
    animal_id: Optional[str] = None
    data: date
    valor: Optional[float] = None
    quantidade: int = 1
    tipo_animal: Optional[str] = None
    observacoes: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EventoCreate(BaseModel):
    tipo: Literal["nascimento", "desmame", "vacinacao", "pesagem", "tratamento"]
    animal_id: str
    data: date
    detalhes: Optional[str] = ""
    peso: Optional[float] = None
    vacina: Optional[str] = None

class Evento(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tipo: Literal["nascimento", "desmame", "vacinacao", "pesagem", "tratamento"]
    animal_id: str
    data: date
    detalhes: Optional[str] = ""
    peso: Optional[float] = None
    vacina: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DespesaCreate(BaseModel):
    categoria_id: str
    valor: float
    data: date
    descricao: Optional[str] = ""

class Despesa(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    categoria_id: str
    valor: float
    data: date
    descricao: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DashboardStats(BaseModel):
    total_animais: int
    total_ativos: int
    total_vendidos: int
    total_mortos: int
    receitas: float
    despesas: float
    lucro: float
    movimentacoes_mes: List[dict]
    despesas_por_categoria: List[dict]


# ============= HELPERS =============

def serialize_doc(doc):
    if isinstance(doc.get('created_at'), str):
        doc['created_at'] = datetime.fromisoformat(doc['created_at'])
    if isinstance(doc.get('data'), str):
        doc['data'] = date.fromisoformat(doc['data'])
    if isinstance(doc.get('data_nascimento'), str):
        doc['data_nascimento'] = date.fromisoformat(doc['data_nascimento'])
    return doc

def prepare_for_db(doc):
    result = doc.copy()
    if 'created_at' in result and isinstance(result['created_at'], datetime):
        result['created_at'] = result['created_at'].isoformat()
    if 'data' in result and isinstance(result['data'], date):
        result['data'] = result['data'].isoformat()
    if 'data_nascimento' in result and result['data_nascimento'] and isinstance(result['data_nascimento'], date):
        result['data_nascimento'] = result['data_nascimento'].isoformat()
    return result


# ============= CATEGORIAS =============

@api_router.post("/categorias", response_model=Categoria)
async def criar_categoria(input: CategoriaCreate):
    categoria = Categoria(**input.model_dump())
    doc = prepare_for_db(categoria.model_dump())
    await db.categorias.insert_one(doc)
    return categoria

@api_router.get("/categorias", response_model=List[Categoria])
async def listar_categorias():
    docs = await db.categorias.find({}, {"_id": 0}).to_list(1000)
    return [serialize_doc(doc) for doc in docs]

@api_router.delete("/categorias/{categoria_id}")
async def deletar_categoria(categoria_id: str):
    result = await db.categorias.delete_one({"id": categoria_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Categoria nao encontrada")
    return {"message": "Categoria deletada"}

@api_router.put("/categorias/{categoria_id}", response_model=Categoria)
async def atualizar_categoria(categoria_id: str, input: CategoriaCreate):
    result = await db.categorias.update_one({"id": categoria_id}, {"$set": {"nome": input.nome, "cor": input.cor}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Categoria nao encontrada")
    doc = await db.categorias.find_one({"id": categoria_id}, {"_id": 0})
    return serialize_doc(doc)


# ============= ANIMAIS =============

@api_router.post("/animais", response_model=Animal)
async def criar_animal(input: AnimalCreate):
    animal = Animal(**input.model_dump())
    doc = prepare_for_db(animal.model_dump())
    await db.animais.insert_one(doc)
    return animal

@api_router.get("/animais", response_model=List[Animal])
async def listar_animais(status: Optional[str] = None):
    filtro = {}
    if status:
        filtro["status"] = status
    docs = await db.animais.find(filtro, {"_id": 0}).to_list(1000)
    return [serialize_doc(doc) for doc in docs]

@api_router.get("/animais/{animal_id}", response_model=Animal)
async def obter_animal(animal_id: str):
    doc = await db.animais.find_one({"id": animal_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Animal nao encontrado")
    return serialize_doc(doc)

@api_router.put("/animais/{animal_id}", response_model=Animal)
async def atualizar_animal(animal_id: str, input: AnimalCreate):
    update_data = prepare_for_db(input.model_dump())
    result = await db.animais.update_one({"id": animal_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Animal nao encontrado")
    doc = await db.animais.find_one({"id": animal_id}, {"_id": 0})
    return serialize_doc(doc)

@api_router.delete("/animais/{animal_id}")
async def deletar_animal(animal_id: str):
    result = await db.animais.delete_one({"id": animal_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Animal nao encontrado")
    return {"message": "Animal deletado"}


# ============= MOVIMENTACOES =============

@api_router.post("/movimentacoes", response_model=Movimentacao)
async def criar_movimentacao(input: MovimentacaoCreate):
    if input.animal_id and input.tipo == "saida":
        await db.animais.update_one(
            {"id": input.animal_id},
            {"$set": {"status": input.motivo if input.motivo in ["venda", "morte", "perda"] else "inativo"}}
        )
    movimentacao = Movimentacao(**input.model_dump())
    doc = prepare_for_db(movimentacao.model_dump())
    await db.movimentacoes.insert_one(doc)
    return movimentacao

@api_router.get("/movimentacoes", response_model=List[Movimentacao])
async def listar_movimentacoes():
    docs = await db.movimentacoes.find({}, {"_id": 0}).sort("data", -1).to_list(1000)
    return [serialize_doc(doc) for doc in docs]

@api_router.delete("/movimentacoes/{movimentacao_id}")
async def deletar_movimentacao(movimentacao_id: str):
    result = await db.movimentacoes.delete_one({"id": movimentacao_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Movimentacao nao encontrada")
    return {"message": "Movimentacao deletada"}


# ============= EVENTOS =============

@api_router.post("/eventos", response_model=Evento)
async def criar_evento(input: EventoCreate):
    if input.tipo == "pesagem" and input.peso:
        await db.animais.update_one({"id": input.animal_id}, {"$set": {"peso_atual": input.peso}})
    evento = Evento(**input.model_dump())
    doc = prepare_for_db(evento.model_dump())
    await db.eventos.insert_one(doc)
    return evento

@api_router.get("/eventos", response_model=List[Evento])
async def listar_eventos(animal_id: Optional[str] = None):
    filtro = {}
    if animal_id:
        filtro["animal_id"] = animal_id
    docs = await db.eventos.find(filtro, {"_id": 0}).sort("data", -1).to_list(1000)
    return [serialize_doc(doc) for doc in docs]

@api_router.delete("/eventos/{evento_id}")
async def deletar_evento(evento_id: str):
    result = await db.eventos.delete_one({"id": evento_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Evento nao encontrado")
    return {"message": "Evento deletado"}


# ============= DESPESAS =============

@api_router.post("/despesas", response_model=Despesa)
async def criar_despesa(input: DespesaCreate):
    despesa = Despesa(**input.model_dump())
    doc = prepare_for_db(despesa.model_dump())
    await db.despesas.insert_one(doc)
    return despesa

@api_router.get("/despesas", response_model=List[Despesa])
async def listar_despesas():
    docs = await db.despesas.find({}, {"_id": 0}).sort("data", -1).to_list(1000)
    return [serialize_doc(doc) for doc in docs]

@api_router.delete("/despesas/{despesa_id}")
async def deletar_despesa(despesa_id: str):
    result = await db.despesas.delete_one({"id": despesa_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Despesa nao encontrada")
    return {"message": "Despesa deletada"}


# ============= DASHBOARD =============

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def obter_stats():
    animais = await db.animais.find({}, {"_id": 0}).to_list(10000)
    movimentacoes = await db.movimentacoes.find({}, {"_id": 0}).to_list(10000)
    despesas = await db.despesas.find({}, {"_id": 0}).to_list(10000)
    categorias = await db.categorias.find({}, {"_id": 0}).to_list(1000)

    total_animais = len(animais)
    total_ativos = len([a for a in animais if a.get("status") == "ativo"])
    total_vendidos = len([a for a in animais if a.get("status") == "venda"])
    total_mortos = len([a for a in animais if a.get("status") in ["morte", "perda"]])

    receitas = sum(m.get("valor", 0) or 0 for m in movimentacoes if m.get("tipo") == "saida" and m.get("motivo") == "venda")
    total_despesas = sum(d.get("valor", 0) for d in despesas)
    custos_entrada = sum(m.get("valor", 0) or 0 for m in movimentacoes if m.get("tipo") == "entrada")
    total_despesas += custos_entrada
    lucro = receitas - total_despesas

    movimentacoes_mes = {}
    for m in movimentacoes:
        data_str = m.get("data")
        mes = str(data_str)[:7] if data_str else "unknown"
        if mes not in movimentacoes_mes:
            movimentacoes_mes[mes] = {"mes": mes, "receitas": 0, "despesas": 0}
        if m.get("tipo") == "saida" and m.get("motivo") == "venda":
            movimentacoes_mes[mes]["receitas"] += m.get("valor", 0) or 0
        elif m.get("tipo") == "entrada":
            movimentacoes_mes[mes]["despesas"] += m.get("valor", 0) or 0

    for d in despesas:
        data_str = d.get("data")
        mes = str(data_str)[:7] if data_str else "unknown"
        if mes not in movimentacoes_mes:
            movimentacoes_mes[mes] = {"mes": mes, "receitas": 0, "despesas": 0}
        movimentacoes_mes[mes]["despesas"] += d.get("valor", 0)

    categorias_dict = {c["id"]: c["nome"] for c in categorias}
    despesas_por_cat = {}
    for d in despesas:
        cat_id = d.get("categoria_id")
        cat_nome = categorias_dict.get(cat_id, "Outros")
        if cat_nome not in despesas_por_cat:
            despesas_por_cat[cat_nome] = 0
        despesas_por_cat[cat_nome] += d.get("valor", 0)

    return DashboardStats(
        total_animais=total_animais, total_ativos=total_ativos,
        total_vendidos=total_vendidos, total_mortos=total_mortos,
        receitas=receitas, despesas=total_despesas, lucro=lucro,
        movimentacoes_mes=sorted(list(movimentacoes_mes.values()), key=lambda x: x["mes"]),
        despesas_por_categoria=[{"categoria": k, "valor": v} for k, v in despesas_por_cat.items()]
    )


# ============= RELATORIOS =============

@api_router.get("/relatorios/pdf")
async def gerar_relatorio_pdf():
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#4A6741'), spaceAfter=30, alignment=TA_CENTER)
    elements.append(Paragraph("Relatorio de Gestao da Fazenda", title_style))
    elements.append(Spacer(1, 0.3*inch))

    stats_response = await obter_stats()
    stats = stats_response.model_dump()

    data = [
        ['Metrica', 'Valor'],
        ['Total de Animais', str(stats['total_animais'])],
        ['Animais Ativos', str(stats['total_ativos'])],
        ['Total de Vendas', str(stats['total_vendidos'])],
        ['Total de Perdas/Mortes', str(stats['total_mortos'])],
        ['Receitas (R$)', f"{stats['receitas']:.2f}"],
        ['Despesas (R$)', f"{stats['despesas']:.2f}"],
        ['Lucro (R$)', f"{stats['lucro']:.2f}"]
    ]
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A6741')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=relatorio_fazenda.pdf"})

@api_router.get("/relatorios/excel")
async def gerar_relatorio_excel():
    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    stats_response = await obter_stats()
    stats = stats_response.model_dump()
    ws_resumo['A1'] = 'Relatorio de Gestao da Fazenda'
    ws_resumo['A1'].font = Font(size=16, bold=True, color='4A6741')
    ws_resumo.merge_cells('A1:B1')
    ws_resumo['A3'] = 'Metrica'
    ws_resumo['B3'] = 'Valor'
    for cell in ['A3', 'B3']:
        ws_resumo[cell].font = Font(bold=True, color='FFFFFF')
        ws_resumo[cell].fill = PatternFill(start_color='4A6741', end_color='4A6741', fill_type='solid')
    metrics = [
        ['Total de Animais', stats['total_animais']], ['Animais Ativos', stats['total_ativos']],
        ['Total de Vendas', stats['total_vendidos']], ['Total de Perdas/Mortes', stats['total_mortos']],
        ['Receitas (R$)', f"{stats['receitas']:.2f}"], ['Despesas (R$)', f"{stats['despesas']:.2f}"],
        ['Lucro (R$)', f"{stats['lucro']:.2f}"]
    ]
    for idx, metric in enumerate(metrics, start=4):
        ws_resumo[f'A{idx}'] = metric[0]
        ws_resumo[f'B{idx}'] = metric[1]
    animais = await db.animais.find({}, {"_id": 0}).to_list(10000)
    if animais:
        ws_animais = wb.create_sheet("Animais")
        for i, h in enumerate(['Tag', 'Tipo', 'Status', 'Peso Atual'], 1):
            ws_animais.cell(row=1, column=i, value=h).font = Font(bold=True)
        for idx, animal in enumerate(animais, start=2):
            ws_animais[f'A{idx}'] = animal.get('tag', '')
            ws_animais[f'B{idx}'] = animal.get('tipo', '')
            ws_animais[f'C{idx}'] = animal.get('status', '')
            ws_animais[f'D{idx}'] = animal.get('peso_atual', '')
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=relatorio_fazenda.xlsx"})


# ============= APP SETUP =============

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    await db.users.create_index("email", unique=True)
    # Seed admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@fazenda.com").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        admin_user = {
            "id": str(uuid.uuid4()),
            "nome": "Administrador",
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(admin_user)
        logger.info(f"Admin criado: {admin_email}")
    elif not verify_password(admin_password, existing.get("password_hash", "")):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})
        logger.info(f"Senha do admin atualizada")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
